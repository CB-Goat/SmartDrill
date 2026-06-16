from openpyxl import load_workbook

wb = load_workbook(r"d:\Projects\素材\记录\人教版小初各学科知识结构（详细版）4.xlsx")

for sheet_name in wb.sheetnames[:3]:
    sheet = wb[sheet_name]
    print(f"\n{'='*60}")
    print(f"Sheet: {sheet_name}")
    print(f"{'='*60}")
    
    for idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=10, values_only=True), start=1):
        if not any(row):
            continue
        print(f"\n行{idx}:")
        for col_idx, val in enumerate(row[:8], start=1):
            col_names = ['年级', '学期', '单元编号', '单元名称', '知识点', '考点', '题型', '频率']
            if val:
                print(f"  {col_names[col_idx-1]}: {str(val)[:100]}")