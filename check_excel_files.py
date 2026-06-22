from openpyxl import load_workbook

print("=" * 60)
print("知识点文件：")
print("=" * 60)
wb = load_workbook(r"d:\Projects\素材\记录\人教版学科单元知识点汇总_20260617-2.xlsx")
for sheet_name in wb.sheetnames[:2]:
    sheet = wb[sheet_name]
    print(f"\nSheet: {sheet_name}")
    for idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=5, values_only=True), start=1):
        if not any(row):
            continue
        print(f"行{idx}: {[str(v)[:30] if v else None for v in row[:6]]}")

print("\n" + "=" * 60)
print("考点文件：")
print("=" * 60)
wb = load_workbook(r"d:\Projects\素材\记录\人教版学科单元考点汇总_20260617-2.xlsx")
for sheet_name in wb.sheetnames[:2]:
    sheet = wb[sheet_name]
    print(f"\nSheet: {sheet_name}")
    for idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=5, values_only=True), start=1):
        if not any(row):
            continue
        print(f"行{idx}: {[str(v)[:30] if v else None for v in row[:8]]}")