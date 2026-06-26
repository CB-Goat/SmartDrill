from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from sqlalchemy.orm import Session
from app.database import get_db, engine
from app.models.models import (
    User, AdminUser, Recharge, Order, Version, Grade, Subject, Semester, Unit,
    KnowledgePoint, ExamPoint, QuestionType, Difficulty, Question
)
from app.utils.auth import get_current_admin
from typing import List
from sse_starlette.sse import EventSourceResponse
import asyncio
import json
import re
import openpyxl
import io
from sqlalchemy import text

router = APIRouter(prefix="/admin", tags=["管理"])

@router.get("/users")
def get_users(admin: User = Depends(get_current_admin), db: Session = Depends(get_db)):
    users = db.query(User).all()
    return [{"id": u.id, "username": u.username, "phone": u.phone, "points": u.points, "role": u.role.value} for u in users]

@router.get("/recharges")
def get_recharges(admin: User = Depends(get_current_admin), db: Session = Depends(get_db)):
    recharges = db.query(Recharge).order_by(Recharge.created_at.desc()).all()
    return [{"id": r.id, "username": r.user.username, "amount": r.amount, "points": r.points, "created_at": r.created_at} for r in recharges]

@router.get("/orders")
def get_orders(admin: User = Depends(get_current_admin), db: Session = Depends(get_db)):
    orders = db.query(Order).order_by(Order.created_at.desc()).all()
    return [{"id": o.id, "username": o.user.username, "title": o.title, "points": o.points, "created_at": o.created_at} for o in orders]

@router.get("/versions")
def get_versions(admin: User = Depends(get_current_admin), db: Session = Depends(get_db)):
    return db.query(Version).all()

@router.post("/versions")
def save_version(data: dict, admin: User = Depends(get_current_admin), db: Session = Depends(get_db)):
    try:
        v = Version(name=data.get('name', ''))
        db.add(v)
        db.commit()
        db.refresh(v)
        return {"id": v.id, "name": v.name}
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"创建版本失败: {str(e)}")

@router.delete("/versions/{id}")
def delete_version(id: int, admin: User = Depends(get_current_admin), db: Session = Depends(get_db)):
    v = db.query(Version).filter(Version.id == id).first()
    if not v:
        raise HTTPException(status_code=404, detail="记录不存在")
    db.delete(v)
    db.commit()
    return {"message": "删除成功"}

@router.get("/grades")
def get_grades(version_id: int = None, admin: User = Depends(get_current_admin), db: Session = Depends(get_db)):
    query = db.query(Grade)
    if version_id:
        query = query.filter(Grade.version_id == version_id)
    return query.all()

@router.post("/grades")
def save_grade(data: dict, admin: User = Depends(get_current_admin), db: Session = Depends(get_db)):
    g = Grade(**data)
    db.add(g)
    db.commit()
    db.refresh(g)
    return {"id": g.id, "name": g.name, "version_id": g.version_id}

@router.delete("/grades/{id}")
def delete_grade(id: int, admin: User = Depends(get_current_admin), db: Session = Depends(get_db)):
    g = db.query(Grade).filter(Grade.id == id).first()
    if not g:
        raise HTTPException(status_code=404, detail="记录不存在")
    db.delete(g)
    db.commit()
    return {"message": "删除成功"}

@router.get("/subjects")
def get_subjects(grade_id: int = None, admin: User = Depends(get_current_admin), db: Session = Depends(get_db)):
    query = db.query(Subject)
    if grade_id:
        query = query.filter(Subject.grade_id == grade_id)
    return query.all()

@router.post("/subjects")
def save_subject(data: dict, admin: User = Depends(get_current_admin), db: Session = Depends(get_db)):
    s = Subject(**data)
    db.add(s)
    db.commit()
    db.refresh(s)
    return {"id": s.id, "name": s.name, "grade_id": s.grade_id}

@router.delete("/subjects/{id}")
def delete_subject(id: int, admin: User = Depends(get_current_admin), db: Session = Depends(get_db)):
    s = db.query(Subject).filter(Subject.id == id).first()
    if not s:
        raise HTTPException(status_code=404, detail="记录不存在")
    db.delete(s)
    db.commit()
    return {"message": "删除成功"}

@router.get("/semesters")
def get_semesters(subject_id: int = None, admin: User = Depends(get_current_admin), db: Session = Depends(get_db)):
    query = db.query(Semester)
    if subject_id:
        query = query.filter(Semester.subject_id == subject_id)
    return query.all()

@router.post("/semesters")
def save_semester(data: dict, admin: User = Depends(get_current_admin), db: Session = Depends(get_db)):
    s = Semester(**data)
    db.add(s)
    db.commit()
    db.refresh(s)
    return {"id": s.id, "name": s.name, "subject_id": s.subject_id}

@router.delete("/semesters/{id}")
def delete_semester(id: int, admin: User = Depends(get_current_admin), db: Session = Depends(get_db)):
    s = db.query(Semester).filter(Semester.id == id).first()
    if not s:
        raise HTTPException(status_code=404, detail="记录不存在")
    db.delete(s)
    db.commit()
    return {"message": "删除成功"}

@router.get("/units")
def get_units(semester_id: int = None, admin: User = Depends(get_current_admin), db: Session = Depends(get_db)):
    query = db.query(Unit)
    if semester_id:
        query = query.filter(Unit.semester_id == semester_id)
    return query.all()

@router.post("/units")
def save_unit(data: dict, admin: User = Depends(get_current_admin), db: Session = Depends(get_db)):
    u = Unit(**data)
    db.add(u)
    db.commit()
    db.refresh(u)
    return {"id": u.id, "name": u.name, "semester_id": u.semester_id}

@router.delete("/units/{id}")
def delete_unit(id: int, admin: User = Depends(get_current_admin), db: Session = Depends(get_db)):
    u = db.query(Unit).filter(Unit.id == id).first()
    if not u:
        raise HTTPException(status_code=404, detail="记录不存在")
    db.delete(u)
    db.commit()
    return {"message": "删除成功"}

@router.get("/units/{id}/knowledge-json")
def get_unit_knowledge_json(id: int, admin: User = Depends(get_current_admin), db: Session = Depends(get_db)):
    u = db.query(Unit).filter(Unit.id == id).first()
    if not u:
        raise HTTPException(status_code=404, detail="单元不存在")
    return u.unit_knowledge_json or {}

@router.put("/units/{id}/knowledge-json")
def update_unit_knowledge_json(id: int, data: dict, admin: User = Depends(get_current_admin), db: Session = Depends(get_db)):
    u = db.query(Unit).filter(Unit.id == id).first()
    if not u:
        raise HTTPException(status_code=404, detail="单元不存在")
    u.unit_knowledge_json = data
    db.commit()
    return {"message": "保存成功"}

@router.get("/knowledge")
def get_knowledge(unit_id: int = None, admin: User = Depends(get_current_admin), db: Session = Depends(get_db)):
    query = db.query(KnowledgePoint)
    if unit_id:
        query = query.filter(KnowledgePoint.unit_id == unit_id)
    return query.all()

@router.post("/knowledge")
def save_knowledge(data: dict, admin: User = Depends(get_current_admin), db: Session = Depends(get_db)):
    if data.get('id'):
        kp = db.query(KnowledgePoint).filter(KnowledgePoint.id == data['id']).first()
        if kp:
            for key, value in data.items():
                setattr(kp, key, value)
            db.commit()
            return {"message": "更新成功"}
    kp = KnowledgePoint(**{k: v for k, v in data.items() if k != 'id'})
    db.add(kp)
    db.commit()
    db.refresh(kp)
    return {"id": kp.id, "message": "保存成功"}

@router.get("/exam-points")
def get_exam_points(knowledge_point_id: int = None, admin: User = Depends(get_current_admin), db: Session = Depends(get_db)):
    query = db.query(ExamPoint)
    if knowledge_point_id:
        query = query.filter(ExamPoint.knowledge_point_id == knowledge_point_id)
    return query.all()

@router.post("/exam-points")
def save_exam_point(data: dict, admin: User = Depends(get_current_admin), db: Session = Depends(get_db)):
    if data.get('id'):
        ep = db.query(ExamPoint).filter(ExamPoint.id == data['id']).first()
        if ep:
            for key, value in data.items():
                setattr(ep, key, value)
            db.commit()
            return {"message": "更新成功"}
    ep = ExamPoint(**{k: v for k, v in data.items() if k != 'id'})
    db.add(ep)
    db.commit()
    db.refresh(ep)
    return {"id": ep.id, "message": "保存成功"}

@router.get("/question-types")
def get_question_types(admin: User = Depends(get_current_admin), db: Session = Depends(get_db)):
    return db.query(QuestionType).all()

@router.get("/difficulties")
def get_difficulties(admin: User = Depends(get_current_admin), db: Session = Depends(get_db)):
    return db.query(Difficulty).all()

@router.get("/questions")
def get_questions(unit_id: int = None, admin: User = Depends(get_current_admin), db: Session = Depends(get_db)):
    query = db.query(Question)
    if unit_id:
        query = query.filter(Question.unit_id == unit_id)
    questions = query.all()
    result = []
    for q in questions:
        result.append({
            "id": q.id,
            "content": q.content,
            "answer": q.answer,
            "analysis": q.analysis,
            "question_type": q.question_type_obj.name if q.question_type_obj else None,
            "difficulty": q.difficulty_obj.name if q.difficulty_obj else None,
            "exam_point_title": q.exam_point.title if q.exam_point else None,
            "question_type_id": q.question_type_id,
            "difficulty_id": q.difficulty_id,
            "unit_id": q.unit_id,
            "knowledge_point_id": q.knowledge_point_id,
            "exam_point_id": q.exam_point_id
        })
    return result

@router.post("/questions")
def save_question(data: dict, admin: User = Depends(get_current_admin), db: Session = Depends(get_db)):
    if data.get('id'):
        q = db.query(Question).filter(Question.id == data['id']).first()
        if q:
            for key, value in data.items():
                setattr(q, key, value)
            db.commit()
            return {"message": "更新成功"}
    q = Question(**{k: v for k, v in data.items() if k != 'id'})
    db.add(q)
    db.commit()
    db.refresh(q)
    return {"id": q.id, "message": "保存成功"}


def _ensure_knowledge_column():
    try:
        with engine.connect() as conn:
            result = conn.execute(text("SHOW COLUMNS FROM units LIKE 'unit_knowledge_json'"))
            if not result.fetchone():
                conn.execute(text("ALTER TABLE units ADD COLUMN unit_knowledge_json JSON"))
                conn.commit()
    except Exception:
        try:
            with engine.connect() as conn:
                conn.execute(text("ALTER TABLE units ADD COLUMN unit_knowledge_json JSON"))
                conn.commit()
        except Exception:
            pass


def _find_unit(db, subject_name, grade_name, semester_name, unit_number):
    version = db.query(Version).filter(Version.name.like('%人教%')).first()
    if not version:
        version = db.query(Version).first()
    if not version:
        return None
    grade = db.query(Grade).filter(Grade.version_id == version.id, Grade.name == grade_name).first()
    if not grade:
        return None
    subject = db.query(Subject).filter(Subject.grade_id == grade.id, Subject.name == subject_name).first()
    if not subject:
        return None
    semester = db.query(Semester).filter(Semester.subject_id == subject.id, Semester.name == semester_name).first()
    if not semester:
        return None
    unit = db.query(Unit).filter(Unit.semester_id == semester.id, Unit.unit_number == unit_number).first()
    return unit


def _parse_array(text):
    if not text:
        return []
    lines = [l.strip() for l in str(text).split('\n') if l.strip()]
    items = []
    for line in lines:
        cleaned = re.sub(r'^\d+[.、．)]\s*', '', line)
        cleaned = re.sub(r'^[①②③④⑤⑥⑦⑧⑨⑩⑪⑫⑬⑭⑮⑯⑰⑱⑲⑳]\s*', '', cleaned)
        if cleaned:
            items.append(cleaned)
    return items


def _sync_knowledge_points(db, unit_id, knowledge_data):
    if not knowledge_data or not isinstance(knowledge_data, dict):
        return 0
    kp_sources = [
        ('核心概念', knowledge_data.get('core_concepts', [])),
        ('重点知识', knowledge_data.get('key_knowledge', [])),
        ('难点解析', knowledge_data.get('difficult_analysis', [])),
        ('易混辨析', knowledge_data.get('confuse_distinction', [])),
    ]
    all_points = []
    for category, items in kp_sources:
        if not isinstance(items, list):
            continue
        for item in items:
            if not isinstance(item, str):
                item = str(item) if item else ''
            if not item:
                continue
            title = f'[{category}] {item[:50]}'
            if len(item) > 50:
                title += '...'
            all_points.append({'title': title, 'content': item})
    if not all_points:
        return 0
    db.query(KnowledgePoint).filter(KnowledgePoint.unit_id == unit_id).delete()
    for point in all_points:
        kp = KnowledgePoint(unit_id=unit_id, title=point['title'], content=point['content'])
        db.add(kp)
    return len(all_points)


@router.post("/knowledge/import-8modules")
async def import_knowledge_8modules(
    file: UploadFile = File(...),
    admin: AdminUser = Depends(get_current_admin),
    db: Session = Depends(get_db)
):
    file_content = await file.read()
    file_name = file.filename or ''

    async def event_generator():
        try:
            _ensure_knowledge_column()
            yield {"event": "message", "data": json.dumps({
                "type": "progress",
                "current": 0,
                "total": 0,
                "status": "info",
                "message": f"读取Excel文件 {file_name}..."
            }, ensure_ascii=False)}
            await asyncio.sleep(0.1)

            wb = openpyxl.load_workbook(io.BytesIO(file_content))
            ws = wb.active
            total_rows = ws.max_row - 1

            yield {"event": "message", "data": json.dumps({
                "type": "progress",
                "current": 0,
                "total": total_rows,
                "status": "info",
                "message": f"共 {total_rows} 个单元，开始匹配导入..."
            }, ensure_ascii=False)}
            await asyncio.sleep(0.1)

            matched = 0
            not_matched = 0
            total_kp = 0
            errors = []

            for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=1):
                try:
                    subject = row[0]
                    grade = row[1]
                    semester = row[2]
                    unit_num = row[3]
                    unit_name = row[5]
                    json_content = row[15] if len(row) > 15 else None

                    if not subject or not unit_name:
                        continue

                    knowledge_data = None
                    if json_content:
                        try:
                            parsed = json.loads(str(json_content))
                            if isinstance(parsed, dict):
                                knowledge_data = parsed
                        except Exception:
                            pass

                    if not knowledge_data:
                        unit_topic = row[6] if len(row) > 6 else ''
                        unit_overview = row[7] if len(row) > 7 else ''
                        knowledge_frame = row[8] if len(row) > 8 else ''
                        core_text = row[9] if len(row) > 9 else ''
                        key_text = row[10] if len(row) > 10 else ''
                        diff_text = row[11] if len(row) > 11 else ''
                        confuse_text = row[12] if len(row) > 12 else ''
                        example_text = row[13] if len(row) > 13 else ''
                        typical_example = {}
                        if example_text:
                            try:
                                parsed_example = json.loads(str(example_text))
                                if isinstance(parsed_example, dict):
                                    typical_example = parsed_example
                                else:
                                    typical_example = {'stem': str(example_text)[:500]}
                            except Exception:
                                typical_example = {'stem': str(example_text)[:500]}
                        knowledge_data = {
                            'unit_topic': str(unit_topic or ''),
                            'unit_overview': str(unit_overview or ''),
                            'knowledge_frame': str(knowledge_frame or ''),
                            'core_concepts': _parse_array(core_text),
                            'key_knowledge': _parse_array(key_text),
                            'difficult_analysis': _parse_array(diff_text),
                            'confuse_distinction': _parse_array(confuse_text),
                            'typical_example': typical_example
                        }

                    unit = _find_unit(db, subject, grade, semester, unit_num)

                    status = "success"
                    msg = ""
                    if unit:
                        matched += 1
                        unit.unit_knowledge_json = knowledge_data
                        try:
                            kp_count = _sync_knowledge_points(db, unit.id, knowledge_data)
                            total_kp += kp_count
                        except Exception as kp_e:
                            errors.append(f'第{idx}行知识点同步失败: {str(kp_e)}')
                        msg = f"✓ {subject}/{grade}/第{unit_num}单元 {unit_name}"
                    else:
                        not_matched += 1
                        status = "error"
                        msg = f"✗ 未匹配: {subject}/{grade}/{semester}/第{unit_num}单元 {unit_name}"
                        errors.append(msg)

                    if idx % 5 == 0 or idx == total_rows:
                        db.commit()
                        yield {"event": "message", "data": json.dumps({
                            "type": "progress",
                            "current": idx,
                            "total": total_rows,
                            "status": status,
                            "message": msg
                        }, ensure_ascii=False)}
                        await asyncio.sleep(0.02)
                    elif idx % 1 == 0:
                        yield {"event": "message", "data": json.dumps({
                            "type": "progress",
                            "current": idx,
                            "total": total_rows,
                            "status": status,
                            "message": msg
                        }, ensure_ascii=False)}
                except Exception as row_e:
                    db.rollback()
                    err_msg = f"第{idx}行处理失败: {str(row_e)}"
                    errors.append(err_msg)
                    yield {"event": "message", "data": json.dumps({
                        "type": "progress",
                        "current": idx,
                        "total": total_rows,
                        "status": "error",
                        "message": err_msg
                    }, ensure_ascii=False)}
                    await asyncio.sleep(0.01)

            yield {"event": "message", "data": json.dumps({
                "type": "complete",
                "current": total_rows,
                "total": total_rows,
                "status": "success",
                "message": f"导入完成！匹配 {matched} 个单元，未匹配 {not_matched} 个，同步知识点 {total_kp} 条",
                "matched": matched,
                "not_matched": not_matched,
                "total_kp": total_kp,
                "errors": errors[:20]
            }, ensure_ascii=False)}

        except Exception as e:
            import traceback
            yield {"event": "message", "data": json.dumps({
                "type": "error",
                "current": 0,
                "total": 0,
                "status": "error",
                "message": f"导入失败: {str(e)}",
                "detail": traceback.format_exc()
            }, ensure_ascii=False)}

    return EventSourceResponse(event_generator())


@router.post("/clear-knowledge-8modules")
def clear_knowledge_8modules(
    data: dict,
    admin: AdminUser = Depends(get_current_admin),
    db: Session = Depends(get_db)
):
    version_id = data.get('version_id')
    grade_id = data.get('grade_id')
    subject_id = data.get('subject_id')
    semester_id = data.get('semester_id')

    query = db.query(Unit)
    if semester_id:
        query = query.filter(Unit.semester_id == semester_id)
    elif subject_id:
        semester_ids = [s.id for s in db.query(Semester).filter(Semester.subject_id == subject_id).all()]
        query = query.filter(Unit.semester_id.in_(semester_ids))
    elif grade_id:
        subject_ids = [s.id for s in db.query(Subject).filter(Subject.grade_id == grade_id).all()]
        semester_ids = [s.id for s in db.query(Semester).filter(Semester.subject_id.in_(subject_ids)).all()]
        query = query.filter(Unit.semester_id.in_(semester_ids))
    elif version_id:
        grade_ids = [g.id for g in db.query(Grade).filter(Grade.version_id == version_id).all()]
        subject_ids = [s.id for s in db.query(Subject).filter(Subject.grade_id.in_(grade_ids)).all()]
        semester_ids = [s.id for s in db.query(Semester).filter(Semester.subject_id.in_(subject_ids)).all()]
        query = query.filter(Unit.semester_id.in_(semester_ids))

    units = query.all()
    cleared_count = 0
    for unit in units:
        if unit.unit_knowledge_json:
            unit.unit_knowledge_json = None
            cleared_count += 1

    db.commit()
    return {"message": f"清除完成，共清除{cleared_count}个单元的8模块知识"}